from io import BytesIO

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def send_excel(buffer, filename):
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def export_document(title, header_fields, detail_columns, detail_rows):
    """單據式匯出：合併標題列 + 單頭資訊 + 明細表格。

    header_fields: [(label, value), ...]
    detail_columns: [(label, key), ...]
    detail_rows: [dict, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    n_cols = max(2, len(detail_columns))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    row = 3
    for label, value in header_fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    header_row = row
    for col_idx, (label, _key) in enumerate(detail_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    for r, item in enumerate(detail_rows, start=header_row + 1):
        for col_idx, (_label, key) in enumerate(detail_columns, start=1):
            cell = ws.cell(row=r, column=col_idx, value=item.get(key))
            cell.border = BORDER

    for col_idx in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer


def export_report(title, columns, rows):
    """平面表格式匯出，用於報表查詢頁。

    columns: [(label, key), ...]
    rows: [dict, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    for col_idx, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    for r, item in enumerate(rows, start=2):
        for col_idx, (_label, key) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=col_idx, value=item.get(key))
            cell.border = BORDER

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer
